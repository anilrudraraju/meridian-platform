"""
Re-ingestion of a reviewed Excel workbook (xlsx_writer.py's output, marked
up by Rich) back into SQLite via review.py.

Matching is by field_value_id (the hidden bookkeeping column in the
Extraction sheet), not by (transaction label, field name) text — a
transaction's display label or a field's display name could in principle
change between versions of the field dictionary, and matching on the
literal id makes re-ingestion correct regardless.

Rule, per the reviewer-status column Rich fills in:
  - "Approved"          -> record_review(..., APPROVE). Final, never
                            re-extracted on a later run (see cli.py).
  - "Rejected"           -> if Rich also changed the Value cell, that's
                            his correction: record_review(..., EDIT,
                            corrected_value=<his value>). If he left Value
                            unchanged, record_review(..., REJECT) — he's
                            flagging it wrong but hasn't supplied a fix;
                            the field is eligible for re-extraction on the
                            next run.
  - "Needs Fix"           -> same edit-vs-no-edit logic as Rejected; the
                            distinction between the two labels is purely
                            for Rich's own bookkeeping (schema.sql only
                            has approve/edit/reject/mark_not_found as
                            actions), both route to the same handling here.
  - blank / "not_attempted" / any other value -> ignored, not yet reviewed.
"""
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from spinoff_research.review import record_review
from spinoff_research.status import ReviewAction


class XlsxReingestionError(Exception):
    pass


@dataclass
class ReingestionResult:
    approved: int = 0
    corrected: int = 0
    rejected_no_fix: int = 0
    skipped_not_reviewed: int = 0
    skipped_no_field_value_id: int = 0
    errors: List[str] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


_HEADERS = [
    "Transaction", "Field", "Category", "Extraction Category", "Value",
    "Status", "Reviewer Status", "Confidence", "Extraction Method",
    "As Of Date", "field_value_id",
]


def reingest_workbook(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    reviewed_by: Optional[str] = None,
) -> ReingestionResult:
    wb = load_workbook(str(xlsx_path), data_only=True)
    if "Extraction" not in wb.sheetnames:
        raise XlsxReingestionError(f"'{xlsx_path}' has no 'Extraction' sheet — not a spinoff_research export")
    ws = wb["Extraction"]

    header_row = [c.value for c in ws[1]]
    try:
        col = {name: header_row.index(name) for name in _HEADERS}
    except ValueError as e:
        raise XlsxReingestionError(f"'{xlsx_path}' Extraction sheet is missing an expected column: {e}")

    result = ReingestionResult()

    for row in ws.iter_rows(min_row=2, values_only=True):
        field_value_id = row[col["field_value_id"]]
        reviewer_status = (row[col["Reviewer Status"]] or "").strip()
        current_value = row[col["Value"]]
        field_name = row[col["Field"]]
        txn_label = row[col["Transaction"]]

        if not reviewer_status:
            result.skipped_not_reviewed += 1
            continue

        if not field_value_id:
            result.skipped_no_field_value_id += 1
            result.errors.append(
                f"'{txn_label}' / '{field_name}': marked '{reviewer_status}' but has no field_value_id "
                f"(field was never extracted — nothing to review yet)."
            )
            continue

        original_row = conn.execute(
            "SELECT * FROM field_values WHERE field_value_id = ?", (int(field_value_id),)
        ).fetchone()
        if original_row is None:
            result.errors.append(f"'{txn_label}' / '{field_name}': field_value_id {field_value_id} not found in database.")
            continue

        original_value = original_row["raw_value"]
        value_changed = (
            current_value is not None
            and str(current_value).strip() != (original_value or "").strip()
        )

        try:
            if reviewer_status == "Approved":
                record_review(conn, int(field_value_id), ReviewAction.APPROVE, reviewed_by=reviewed_by)
                result.approved += 1
            elif reviewer_status in ("Rejected", "Needs Fix"):
                if value_changed:
                    record_review(
                        conn, int(field_value_id), ReviewAction.EDIT,
                        corrected_value=str(current_value), reviewed_by=reviewed_by,
                        reviewer_notes=f"Corrected via Excel round-trip (was: {original_value!r})",
                    )
                    result.corrected += 1
                else:
                    record_review(
                        conn, int(field_value_id), ReviewAction.REJECT, reviewed_by=reviewed_by,
                        reviewer_notes=f"Marked '{reviewer_status}' via Excel round-trip, no correction supplied",
                    )
                    result.rejected_no_fix += 1
            else:
                result.errors.append(
                    f"'{txn_label}' / '{field_name}': unrecognized Reviewer Status '{reviewer_status}' — ignored."
                )
        except ValueError as e:
            result.errors.append(f"'{txn_label}' / '{field_name}': {e}")

    return result
