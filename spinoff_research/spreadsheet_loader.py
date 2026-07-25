"""
Loads Rich's historical spin-off workbook and maps it onto the internal
field data dictionary. Read-only: never writes back to the source workbook.

Rich's real workbook ("Stock Spin-off Investing Database.xlsx", single
sheet "2024 US Spin-offs") has project notes, an email address, and a
credential-like string in rows 1-9 above the real header at row 10 — those
rows are never read by this loader. Only data/input/rich_spinoff_database_raw.xlsx
(header + data rows only, already stripped) should be pointed at in practice.

Column mapping is by exact match against FieldDefinition.spreadsheet_column_name
after whitespace normalization (Rich's headers have inconsistent trailing
spaces, e.g. "Parent Company " with a trailing space). Unmapped columns are
reported, not silently dropped, so drift between the sheet and the data
dictionary is always visible.
"""
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from spinoff_research.field_data_dictionary import FIELD_DEFINITIONS


def _normalize_header(h: Optional[str]) -> str:
    return (h or "").strip()


def _match_key(h: Optional[str]) -> str:
    """Case-insensitive matching key — Rich's headers and the data dictionary's
    spreadsheet_column_name disagree on casing in several places (e.g.
    'Parent Ticker' vs 'Parent ticker'); matching should not be case-sensitive,
    but the *display* header (from _normalize_header) is preserved as-is for
    traceability."""
    return _normalize_header(h).lower()


@dataclass
class LoadedCell:
    field_key: Optional[str]        # None if the column couldn't be mapped
    spreadsheet_column_name: str
    raw_value: Any                  # value as read (data_only load — formula RESULT, not formula text)
    is_formula: bool
    formula_text: Optional[str] = None
    is_blank: bool = False
    has_stale_formula_cache: bool = False  # formula present but no cached result (file never recalculated by a spreadsheet app)


@dataclass
class LoadedTransactionRow:
    source_row_number: int          # 1-indexed row in the sheet (post-header)
    sheet_row: int                  # actual openpyxl row number
    cells: Dict[str, LoadedCell] = dc_field(default_factory=dict)  # keyed by field_key ('' for unmapped)

    def get(self, field_key: str) -> Optional[LoadedCell]:
        return self.cells.get(field_key)


@dataclass
class LoadResult:
    rows: List[LoadedTransactionRow]
    unmapped_columns: List[str]        # spreadsheet columns with no matching field_key
    missing_fields: List[str]          # field_keys in the dictionary not found as a column
    header_row_index: int


# Column name -> field_key, built from the data dictionary's spreadsheet_column_name.
# Keyed case-insensitively — see _match_key.
_COLUMN_TO_FIELD_KEY = {
    _match_key(f.spreadsheet_column_name): f.field_key
    for f in FIELD_DEFINITIONS
}

# Rich's real header labels differ slightly from the ones captured in the
# original "Field Source Review" hypothesis sheet used to seed the data
# dictionary (casing, "Tenor" vs "tenure", "Intially" typo, extra "?").
# Map those known variants explicitly rather than fuzzy-matching, so mapping
# stays auditable.
_KNOWN_HEADER_ALIASES = {
    # keys normalized via _match_key() at lookup time; written here in
    # natural case purely for human readability
    "Form 10": "form_10_availability",
    "Parent Company": "parent_company_name",
    "Spun-off Company": "spinoff_company_name",
    "Spin-off Announcement Date": "announcement_date",
    "Beginning of Regular Way Trading": "regular_way_trading_date",
    "Parent Shares Outstanding": "parent_shares_outstanding",
    "Parent Market Cap (BN)": "parent_market_cap",
    "Spin-off Shares Outstanding (MM)": "spinoff_shares_outstanding",
    "Spin-off Market Cap (BN)": "spinoff_market_cap",
    "Spin-off Debt?": "spinoff_debt",
    "Spin-off Debt to EBITDA?": "spinoff_debt_to_ebitda",
    "Did CEO come from Parent?": "ceo_came_from_parent",
    "CEO tenor (years) at parent?": "ceo_tenure_at_parent",
    "Did CFO come from Parent?": "cfo_came_from_parent",
    "CFO Tenor at Parent?": "cfo_tenure_at_parent",
    "One Time Expenses (MM)": "one_time_separation_expenses",
    "Dis-synergy Guidance (MM)": "dis_synergy_guidance",
    "Does the Spin-off Take Place Before/In Conjunction with the Parent Company Acquired?": "spin_preceded_or_with_parent_acquisition",
    "Last Year Sales (BN)": "last_year_sales",
    "Dis-Synergy Guidance as a % of Last Year Sales": "dis_synergy_pct_of_prior_year_sales",
    "Insider Buying within 3 months of Spin-off?": "insider_buying_within_3mo",
    "Cluster Insider Buy with 3 Months?": "cluster_insider_buying_within_3mo",
    "How Many Insiders Bought?": "insider_buyer_count",
    "Revenue Growth Guidance for Current Year": "revenue_growth_guidance_current_year",
    "Intially Paying a Dividend?": "initially_paying_dividend",
    "Initiated a Dividend Within 12 Months?": "dividend_initiated_within_12mo",
    "Preannounced Share Buyback?": "preannounced_buyback",
    "Was the Spin-off Regulatory Driven?": "is_regulatory_driven",
    "Divisions in Parent Company": "parent_division_count",
    "Reversing a Failed Acquisition?": "reversed_failed_acquisition",
    "100% Spin-off?": "is_100_percent_spinoff",
    "TSR Based Incentive Plan": "tsr_based_incentive_plan",
    "Spinoff EBITDA Margin": "spinoff_ebitda_margin",
}


_ALIASES_BY_MATCH_KEY = {_match_key(k): v for k, v in _KNOWN_HEADER_ALIASES.items()}


def _map_column(header: str) -> Optional[str]:
    h = _match_key(header)
    if h in _COLUMN_TO_FIELD_KEY:
        return _COLUMN_TO_FIELD_KEY[h]
    if h in _ALIASES_BY_MATCH_KEY:
        return _ALIASES_BY_MATCH_KEY[h]
    return None


def load_workbook(
    path: Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> LoadResult:
    """
    Load a spin-off transaction table from an .xlsx file.

    path: file with header at `header_row` and data immediately below.
    sheet_name: defaults to the workbook's first/active sheet.
    header_row: 1-indexed row containing column headers. Rich's raw export
        (before stripping) has the real header at row 10 — always point this
        loader at the pre-stripped file (header_row=1) in normal use.

    Never modifies or writes back to the source file.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    wb_formulas = openpyxl.load_workbook(str(path), data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    ws_formulas = wb_formulas[sheet_name] if sheet_name else wb_formulas.active

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    header_map: Dict[int, Optional[str]] = {
        col: _map_column(h) for col, h in enumerate(headers, start=1) if h is not None
    }
    unmapped = sorted({
        _normalize_header(headers[col - 1])
        for col, key in header_map.items()
        if key is None
    })

    rows: List[LoadedTransactionRow] = []
    src_row_num = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue  # fully blank row — stop counting as data but keep scanning in case of gaps
        src_row_num += 1
        loaded = LoadedTransactionRow(source_row_number=src_row_num, sheet_row=r)
        for col, field_key in header_map.items():
            if field_key is None:
                continue
            raw_val = ws.cell(row=r, column=col).value
            formula_val = ws_formulas.cell(row=r, column=col).value
            is_formula = isinstance(formula_val, str) and formula_val.startswith("=")
            # A formula with no cached result (raw_val is None) means the workbook
            # was never recalculated by a spreadsheet application before being
            # read — the value is genuinely unknown, not "blank" in the normal
            # sense. Surface this distinctly so a loader caller doesn't silently
            # treat an un-evaluated formula as a not-found value.
            has_stale_cache = is_formula and raw_val is None
            is_blank = (not is_formula) and (
                raw_val is None or (isinstance(raw_val, str) and not raw_val.strip())
            )
            loaded.cells[field_key] = LoadedCell(
                field_key=field_key,
                spreadsheet_column_name=_normalize_header(headers[col - 1]),
                raw_value=raw_val,
                is_formula=is_formula,
                formula_text=formula_val if is_formula else None,
                is_blank=is_blank,
                has_stale_formula_cache=has_stale_cache,
            )
        rows.append(loaded)

    mapped_keys = {k for k in header_map.values() if k is not None}
    all_dictionary_keys = {f.field_key for f in FIELD_DEFINITIONS}
    missing_fields = sorted(all_dictionary_keys - mapped_keys)

    return LoadResult(
        rows=rows,
        unmapped_columns=unmapped,
        missing_fields=missing_fields,
        header_row_index=header_row,
    )
