"""
Excel export for the extraction pipeline — the "minimal UI" surface for
Rich's offline review, replacing the Streamlit review UI. SQLite
(field_values / field_sources / field_reviews) remains the source of
truth; this module is a read-only view over it plus a run summary,
written to a versioned .xlsx file per CLI invocation.

Workbook layout:
  - "Input"      — the transaction list requested for this run.
  - "Extraction" — one row per (transaction, field): current value,
                    status (color-coded), confidence (data bar), and the
                    source citation as a cell comment. This is the sheet
                    Rich actually marks up.
  - "Run Log"     — what this run did: newly extracted / skipped (why) /
                    carried through approved-and-untouched / re-attempted
                    from a prior round's feedback.
"""
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from spinoff_research.field_data_dictionary import FIELD_BY_KEY, FIELD_DEFINITIONS
from spinoff_research.review_data import load_review_rows
from spinoff_research.repository import get_transaction

# Same 9-state palette as the Streamlit design brief/handoff — kept
# identical on purpose so a future UI (if ever built) and this workbook
# read as the same system, not two different color languages.
_STATUS_FILL = {
    "approved":                   PatternFill("solid", fgColor="E5F4EC"),
    "extracted_high_confidence":  PatternFill("solid", fgColor="E7F3F8"),
    "extracted_uncertain":        PatternFill("solid", fgColor="FBF1E3"),
    "not_found":                  PatternFill("solid", fgColor="ECECEC"),
    "conflicting_sources":        PatternFill("solid", fgColor="FBEAE8"),
    "not_yet_determinable":       PatternFill("solid", fgColor="EEEAFA"),
    "requires_manual_review":     PatternFill("solid", fgColor="E3F3F3"),
    "rejected":                   PatternFill("solid", fgColor="FAEAEA"),
    "manually_entered":           PatternFill("solid", fgColor="EEEEEE"),
}
_STATUS_FONT_COLOR = {
    "approved": "1A7A4C", "extracted_high_confidence": "1A6B8F",
    "extracted_uncertain": "A3631A", "not_found": "6B6B6B",
    "conflicting_sources": "B23A2E", "not_yet_determinable": "5B4B9E",
    "requires_manual_review": "2A7A7A", "rejected": "A32020",
    "manually_entered": "5A5A5A",
}

_STATUS_DROPDOWN_OPTIONS = ["", "Approved", "Rejected", "Needs Fix"]

_EXTRACTION_HEADERS = [
    "Transaction", "Field", "Category", "Extraction Category", "Value",
    "Status", "Reviewer Status", "Confidence", "Extraction Method",
    "As Of Date", "field_value_id",
]
# field_value_id is a hidden bookkeeping column: re-ingestion matches
# Rich's edited rows back to the exact field_values row they came from,
# rather than re-matching by (transaction, field) text, which would break
# if a transaction's display label or field display name ever changes.

_HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


@dataclass
class RunLogEntry:
    transaction_label: str
    field_key: str
    outcome: str          # 'newly_extracted' | 'skipped_approved' | 'reattempted' | 'skipped_no_mechanism'
    detail: str = ""


def _confidence_band_note(confidence: Optional[float]) -> str:
    if confidence is None:
        return ""
    if confidence >= 0.85:
        return "high"
    if confidence >= 0.60:
        return "medium"
    return "low"


def write_workbook(
    conn: sqlite3.Connection,
    transaction_ids: List[int],
    output_path: Path,
    run_log: Optional[List[RunLogEntry]] = None,
) -> Path:
    wb = Workbook()
    _write_input_sheet(wb, conn, transaction_ids)
    _write_extraction_sheet(wb, conn, transaction_ids)
    _write_run_log_sheet(wb, run_log or [])
    del wb["Sheet"]  # openpyxl's default blank sheet

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _write_input_sheet(wb: Workbook, conn: sqlite3.Connection, transaction_ids: List[int]) -> None:
    ws = wb.create_sheet("Input")
    headers = ["Transaction ID", "Parent", "Parent Ticker", "Spinoff", "Spinoff Ticker",
               "Announcement Date", "Distribution Date", "Requested At"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    now = datetime.now().isoformat(timespec="seconds")
    for txn_id in transaction_ids:
        txn = get_transaction(conn, txn_id)
        if txn is None:
            continue
        ws.append([
            txn.transaction_id, txn.parent.name, txn.parent.ticker,
            txn.spinoff.name, txn.spinoff.ticker,
            txn.announcement_date or "", txn.spinoff_date or "", now,
        ])

    widths = [14, 28, 12, 28, 12, 16, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_extraction_sheet(wb: Workbook, conn: sqlite3.Connection, transaction_ids: List[int]) -> None:
    ws = wb.create_sheet("Extraction")
    ws.append(_EXTRACTION_HEADERS)
    for c in range(1, len(_EXTRACTION_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    status_col = _EXTRACTION_HEADERS.index("Status") + 1
    reviewer_status_col = _EXTRACTION_HEADERS.index("Reviewer Status") + 1
    confidence_col = _EXTRACTION_HEADERS.index("Confidence") + 1
    value_col = _EXTRACTION_HEADERS.index("Value") + 1

    row_idx = 1
    for txn_id in transaction_ids:
        txn = get_transaction(conn, txn_id)
        if txn is None:
            continue
        rows = load_review_rows(conn, txn_id)
        rows_by_key = {r.field_key: r for r in rows}

        # Every field the dictionary knows about gets a row, not just
        # ones that were extracted — a blank Value + "not attempted"
        # Status makes gaps visible in the sheet itself, same as the
        # Streamlit design's "locked" rows for not-yet-built fields.
        for field_def in FIELD_DEFINITIONS:
            row = rows_by_key.get(field_def.field_key)
            row_idx += 1
            citation_text = None
            if row is None:
                ws.append([
                    txn.label(), field_def.display_name, field_def.category,
                    field_def.extraction_category.value, "", "not_attempted", "",
                    "", "", "", "",
                ])
            else:
                confidence_pct = row.confidence if row.confidence is not None else None
                ws.append([
                    txn.label(), field_def.display_name, field_def.category,
                    field_def.extraction_category.value, row.raw_value or "",
                    row.status, "", confidence_pct if confidence_pct is not None else "",
                    "", row.as_of_date or "", row.field_value_id,
                ])
                if row.sources:
                    citation_text = "\n\n".join(
                        f"{s.reasoning_summary}" + (f"\n{s.supporting_excerpt}" if s.supporting_excerpt else "")
                        + (f"\n{s.sec_url}" if s.sec_url else "")
                        for s in row.sources
                    )

            status_val = ws.cell(row=row_idx, column=status_col).value
            fill = _STATUS_FILL.get(status_val)
            font_color = _STATUS_FONT_COLOR.get(status_val)
            if fill:
                ws.cell(row=row_idx, column=status_col).fill = fill
            if font_color:
                ws.cell(row=row_idx, column=status_col).font = Font(color=font_color, bold=True)

            if citation_text:
                comment = Comment(citation_text[:2000], "spinoff_research")
                comment.width, comment.height = 350, 150
                ws.cell(row=row_idx, column=value_col).comment = comment

    # Reviewer Status dropdown — the ONLY column Rich is expected to edit
    # besides Value itself.
    dv = DataValidation(type="list", formula1='"' + ",".join(_STATUS_DROPDOWN_OPTIONS[1:]) + '"', allow_blank=True)
    dv.error = "Choose Approved, Rejected, or Needs Fix"
    dv.prompt = "Mark your review decision for this field"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(reviewer_status_col)}2:{get_column_letter(reviewer_status_col)}{row_idx}")

    # Confidence data bar
    if row_idx > 1:
        rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                            color="1A7A4C", showValue=True, minLength=None, maxLength=None)
        ws.conditional_formatting.add(
            f"{get_column_letter(confidence_col)}2:{get_column_letter(confidence_col)}{row_idx}", rule
        )

    widths = {"A": 32, "B": 34, "C": 26, "D": 22, "E": 22, "F": 20, "G": 16, "H": 14, "I": 18, "J": 14, "K": 4}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_EXTRACTION_HEADERS))}{row_idx}"
    # hide the bookkeeping id column rather than delete it — re-ingestion needs it
    ws.column_dimensions[get_column_letter(len(_EXTRACTION_HEADERS))].hidden = True


def _write_run_log_sheet(wb: Workbook, run_log: List[RunLogEntry]) -> None:
    ws = wb.create_sheet("Run Log")
    headers = ["Transaction", "Field", "Outcome", "Detail"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for entry in run_log:
        ws.append([entry.transaction_label, entry.field_key, entry.outcome, entry.detail])
    widths = [32, 34, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
