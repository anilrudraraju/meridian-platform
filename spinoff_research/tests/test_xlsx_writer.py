import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from spinoff_research.db import init_db
from spinoff_research.extraction import ExtractedFieldValue, SourceCitation, persist_field_value
from spinoff_research.models import Company, SpinoffTransaction
from spinoff_research.repository import get_or_create_transaction
from spinoff_research.status import FieldStatus
from spinoff_research.xlsx_writer import write_workbook, RunLogEntry


class TestXlsxWriter(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.conn = init_db(Path(self._tmpdir.name) / "test.db")
        tx = SpinoffTransaction(
            parent=Company(name="Parent Co", ticker="P"),
            spinoff=Company(name="Spinco", ticker="S"),
            announcement_date="2024-01-01", spinoff_date="2024-04-01",
        )
        self.txn = get_or_create_transaction(self.conn, tx)

    def tearDown(self):
        self.conn.close()
        self._tmpdir.cleanup()

    def test_writes_three_sheets(self):
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        self.assertEqual(set(wb.sheetnames), {"Input", "Extraction", "Run Log"})

    def test_extraction_sheet_has_one_row_per_dictionary_field(self):
        from spinoff_research.field_data_dictionary import FIELD_DEFINITIONS
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        self.assertEqual(ws.max_row - 1, len(FIELD_DEFINITIONS))  # -1 for header

    def test_unattempted_field_shows_not_attempted_status(self):
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        status_col = headers.index("Status")
        statuses = {row[status_col] for row in ws.iter_rows(min_row=2, values_only=True)}
        self.assertIn("not_attempted", statuses)

    def test_extracted_field_appears_with_value_and_status(self):
        result = ExtractedFieldValue(
            field_key="form_10_availability", extraction_method="filing_metadata",
            status=FieldStatus.EXTRACTED_HIGH_CONFIDENCE, raw_value="True", confidence=1.0,
        )
        persist_field_value(self.conn, self.txn.transaction_id, result)
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        field_col, value_col, status_col = headers.index("Field"), headers.index("Value"), headers.index("Status")
        found = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[field_col] == "Form 10 availability"]
        self.assertEqual(len(found), 1)
        self.assertEqual(found[0][value_col], "True")
        self.assertEqual(found[0][status_col], "extracted_high_confidence")

    def test_extraction_method_column_populated(self):
        """Regression: this column was previously always blank — see
        review_data.py's load_review_rows fix."""
        result = ExtractedFieldValue(
            field_key="form_10_availability", extraction_method="filing_metadata",
            status=FieldStatus.EXTRACTED_HIGH_CONFIDENCE, raw_value="True",
        )
        persist_field_value(self.conn, self.txn.transaction_id, result)
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        field_col, method_col = headers.index("Field"), headers.index("Extraction Method")
        found = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[field_col] == "Form 10 availability"]
        self.assertEqual(found[0][method_col], "filing_metadata")

    def test_extraction_method_includes_model_for_ai_assisted_fields(self):
        result = ExtractedFieldValue(
            field_key="ceo_came_from_parent", extraction_method="ai_assisted",
            status=FieldStatus.EXTRACTED_HIGH_CONFIDENCE, raw_value="True", model_used="claude-haiku-4-5",
        )
        persist_field_value(self.conn, self.txn.transaction_id, result)
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        field_col, method_col = headers.index("Field"), headers.index("Extraction Method")
        found = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[field_col] == "Did the CEO come from the parent?"]
        self.assertEqual(found[0][method_col], "ai_assisted (claude-haiku-4-5)")

    def test_source_citation_becomes_cell_comment(self):
        result = ExtractedFieldValue(
            field_key="form_10_availability", extraction_method="filing_metadata",
            status=FieldStatus.EXTRACTED_HIGH_CONFIDENCE, raw_value="True",
            sources=[SourceCitation(reasoning_summary="Found via EDGAR", supporting_excerpt="10-12B filed 2024-02-15")],
        )
        persist_field_value(self.conn, self.txn.transaction_id, result)
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        value_col = headers.index("Value") + 1
        found_comment = False
        for row in ws.iter_rows(min_row=2):
            if row[value_col - 1].value == "True" and row[value_col - 1].comment:
                self.assertIn("Found via EDGAR", row[value_col - 1].comment.text)
                found_comment = True
        self.assertTrue(found_comment)

    def test_field_value_id_column_present_and_hidden(self):
        result = ExtractedFieldValue(field_key="form_10_availability", extraction_method="filing_metadata", status=FieldStatus.NOT_FOUND)
        persist_field_value(self.conn, self.txn.transaction_id, result)
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        self.assertIn("field_value_id", headers)
        id_col_letter = ws.cell(row=1, column=headers.index("field_value_id") + 1).column_letter
        self.assertTrue(ws.column_dimensions[id_col_letter].hidden)

    def test_reviewer_status_has_dropdown_validation(self):
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        self.assertEqual(len(ws.data_validations.dataValidation), 1)
        dv = ws.data_validations.dataValidation[0]
        self.assertIn("Approved", dv.formula1)
        self.assertIn("Rejected", dv.formula1)
        self.assertIn("Needs Fix", dv.formula1)

    def test_confidence_has_data_bar_rule(self):
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Extraction"]
        self.assertGreaterEqual(len(ws.conditional_formatting._cf_rules), 1)

    def test_input_sheet_reflects_requested_transaction(self):
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Input"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][1], "Parent Co")
        self.assertEqual(rows[0][2], "P")

    def test_run_log_sheet_reflects_entries(self):
        log = [RunLogEntry("Spinco spun off from Parent Co", "form_10_availability", "newly_extracted", "status=extracted_high_confidence")]
        out = write_workbook(self.conn, [self.txn.transaction_id], Path(self._tmpdir.name) / "out.xlsx", run_log=log)
        wb = load_workbook(out)
        ws = wb["Run Log"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][2], "newly_extracted")

    def test_skips_missing_transaction_id_gracefully(self):
        out = write_workbook(self.conn, [self.txn.transaction_id, 999999], Path(self._tmpdir.name) / "out.xlsx")
        wb = load_workbook(out)
        ws = wb["Input"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)  # only the real transaction, 999999 silently skipped


if __name__ == "__main__":
    unittest.main()
