import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from spinoff_research.db import init_db
from spinoff_research.extraction import ExtractedFieldValue, persist_field_value, get_current_field_value
from spinoff_research.models import Company, SpinoffTransaction
from spinoff_research.repository import get_or_create_transaction
from spinoff_research.status import FieldStatus
from spinoff_research.xlsx_writer import write_workbook
from spinoff_research.xlsx_reader import reingest_workbook, XlsxReingestionError


class TestXlsxReader(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.conn = init_db(Path(self._tmpdir.name) / "test.db")
        tx = SpinoffTransaction(
            parent=Company(name="Parent Co", ticker="P"),
            spinoff=Company(name="Spinco", ticker="S"),
            announcement_date="2024-01-01", spinoff_date="2024-04-01",
        )
        self.txn = get_or_create_transaction(self.conn, tx)
        persist_field_value(self.conn, self.txn.transaction_id, ExtractedFieldValue(
            field_key="form_10_availability", extraction_method="filing_metadata",
            status=FieldStatus.EXTRACTED_HIGH_CONFIDENCE, raw_value="True",
        ))
        persist_field_value(self.conn, self.txn.transaction_id, ExtractedFieldValue(
            field_key="parent_sector", extraction_method="market_data",
            status=FieldStatus.EXTRACTED_UNCERTAIN, raw_value="Diagnostics & Research",
        ))
        self.xlsx_path = Path(self._tmpdir.name) / "export.xlsx"
        write_workbook(self.conn, [self.txn.transaction_id], self.xlsx_path)

    def tearDown(self):
        self.conn.close()
        self._tmpdir.cleanup()

    def _set_reviewer_status(self, path, field_display_name, reviewer_status, new_value=None):
        wb = load_workbook(path)
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        field_col = headers.index("Field")
        status_col = headers.index("Reviewer Status") + 1
        value_col = headers.index("Value") + 1
        for row in ws.iter_rows(min_row=2):
            if row[field_col].value == field_display_name:
                row[status_col - 1].value = reviewer_status
                if new_value is not None:
                    row[value_col - 1].value = new_value
        wb.save(path)

    def test_approve_without_value_change(self):
        self._set_reviewer_status(self.xlsx_path, "Form 10 availability", "Approved")
        result = reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        self.assertEqual(result.approved, 1)
        current = get_current_field_value(self.conn, self.txn.transaction_id, "form_10_availability")
        self.assertEqual(current["status"], "approved")
        self.assertEqual(current["raw_value"], "True")  # unchanged

    def test_needs_fix_with_value_change_becomes_correction(self):
        self._set_reviewer_status(self.xlsx_path, "Parent sector", "Needs Fix", new_value="Health Care")
        result = reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        self.assertEqual(result.corrected, 1)
        current = get_current_field_value(self.conn, self.txn.transaction_id, "parent_sector")
        self.assertEqual(current["status"], "manually_entered")
        self.assertEqual(current["raw_value"], "Health Care")

    def test_rejected_without_value_change_is_reject_only(self):
        self._set_reviewer_status(self.xlsx_path, "Parent sector", "Rejected")
        result = reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        self.assertEqual(result.rejected_no_fix, 1)
        self.assertEqual(result.corrected, 0)

    def test_blank_reviewer_status_is_skipped(self):
        result = reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        self.assertEqual(result.approved, 0)
        self.assertEqual(result.corrected, 0)
        self.assertGreater(result.skipped_not_reviewed, 0)

    def test_unrecognized_reviewer_status_is_reported_as_error(self):
        self._set_reviewer_status(self.xlsx_path, "Form 10 availability", "Maybe??")
        result = reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        self.assertEqual(len(result.errors), 1)
        self.assertIn("Maybe??", result.errors[0])

    def test_missing_extraction_sheet_raises(self):
        from openpyxl import Workbook
        bad = Path(self._tmpdir.name) / "bad.xlsx"
        wb = Workbook()
        wb.save(bad)
        with self.assertRaises(XlsxReingestionError):
            reingest_workbook(self.conn, bad)

    def test_reingest_is_idempotent_when_rerun_without_new_edits(self):
        """Running reingest twice on the SAME reviewed file should not
        double-apply — the second pass re-approves an already-approved
        row, which is a valid no-op (approve is idempotent on schema.sql's
        own design), not a duplicate correction."""
        self._set_reviewer_status(self.xlsx_path, "Form 10 availability", "Approved")
        reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        reingest_workbook(self.conn, self.xlsx_path, reviewed_by="Rich")
        current = get_current_field_value(self.conn, self.txn.transaction_id, "form_10_availability")
        self.assertEqual(current["status"], "approved")


if __name__ == "__main__":
    unittest.main()
